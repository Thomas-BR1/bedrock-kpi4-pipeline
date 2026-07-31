#!/usr/bin/env python3
"""
Combined-file helper for the KPI 4 skill.

Adds a new dated snapshot tab to the "Bedrock Restoration Memberships Combined File"
.xlsx and updates the "Name Tracking" tab per the tracking rules:

  - Compare the new snapshot to the current Name Tracking state.
  - Any (Display Name, Plan) that was Active in Name Tracking but is missing/non-Active
    in the new snapshot -> mark Cancelled, End Date = pull_date.
  - Any (Display Name, Plan) that is Active in the new snapshot but has never appeared
    in Name Tracking -> append as Active with First Seen (Pull) = pull_date.

Returns a diff summary the caller can use in the KPI 4 report.
"""
import csv
import datetime
import os
import re
from copy import copy

try:
    from openpyxl import load_workbook
except ImportError as e:
    raise SystemExit("openpyxl is required. Install with: pip install openpyxl") from e


NT_HEADER = ["Display Name", "Plan", "Start Date", "Status", "End Date", "First Seen (Pull)"]
SNAPSHOT_HEADER = [
    "Plan", "Start Date", "End Date", "Status",
    "First Name", "Last Name", "Display Name",
    "Mobile Number", "Home Number", "Email", "Company", "ID",
    "Address Street Line 1", "Address Street Line 2",
    "Address City", "Address State", "Address Postal Code",
    "Address Billing?", "Address Notes",
]


def extract_pull_date(csv_path, override=None):
    """
    Determine the pull_date (YYYY-MM-DD) for this snapshot:
      1. If override is given, use it.
      2. Else look for a YYYY-MM-DD pattern in the CSV filename.
      3. Else fall back to today's date.
    """
    if override:
        return override
    m = re.search(r"(\d{4}-\d{2}-\d{2})", os.path.basename(csv_path))
    if m:
        return m.group(1)
    return datetime.date.today().isoformat()


def read_csv_rows(csv_path):
    with open(csv_path, newline="", encoding="utf-8") as f:
        return list(csv.DictReader(f))


def _norm(s):
    if s is None:
        return ""
    # openpyxl may return datetime/date for date-typed cells; normalize to ISO string.
    if isinstance(s, datetime.datetime):
        return s.date().isoformat()
    if isinstance(s, datetime.date):
        return s.isoformat()
    return str(s).strip()


def build_current_active(csv_rows, excluded_plans=None):
    """
    Return dict keyed by (Display Name, Plan) -> {start_date, ...} for Active rows only.
    Rows whose plan is in `excluded_plans` are skipped so they don't participate in the diff.
    """
    excluded_plans = excluded_plans or set()
    out = {}
    for r in csv_rows:
        if _norm(r.get("Status")) != "Active":
            continue
        name = _norm(r.get("Display Name"))
        plan = _norm(r.get("Plan"))
        if not name or not plan:
            continue
        if plan in excluded_plans:
            continue
        key = (name, plan)
        start = _norm(r.get("Start Date"))
        if key not in out or (start and start < out[key]["start_date"]):
            out[key] = {"start_date": start, "row": r}
    return out


def read_name_tracking(ws):
    """
    Return (header, rows) where rows is a list of dicts with row_index preserved.
    row_index is 1-based openpyxl row number for later updates.
    """
    header = [ _norm(ws.cell(row=1, column=c).value) for c in range(1, ws.max_column + 1) ]
    # Basic sanity check
    for expected in NT_HEADER:
        if expected not in header:
            raise RuntimeError(
                f"Name Tracking header missing column '{expected}'. Got: {header}"
            )
    col_index = {name: header.index(name) + 1 for name in NT_HEADER}

    rows = []
    for r in range(2, ws.max_row + 1):
        rec = {name: _norm(ws.cell(row=r, column=col_index[name]).value) for name in NT_HEADER}
        if not rec["Display Name"] and not rec["Plan"]:
            continue
        rec["_row"] = r
        rows.append(rec)
    return header, col_index, rows


def read_all_name_tracking(combined_xlsx_path, excluded_plans=None):
    """
    Return every row of the Name Tracking tab as dicts (both Active and Cancelled).
    Excludes plans in `excluded_plans`.
    """
    excluded_plans = excluded_plans or set()
    wb = load_workbook(combined_xlsx_path, data_only=True)
    if "Name Tracking" not in wb.sheetnames:
        return []
    ws = wb["Name Tracking"]
    _, col_index, nt_rows = read_name_tracking(ws)
    out = []
    for r in nt_rows:
        if r["Plan"] in excluded_plans:
            continue
        out.append({
            "display_name": r["Display Name"],
            "plan": r["Plan"],
            "start_date": r["Start Date"],
            "end_date": r["End Date"],
            "status": r["Status"],
        })
    return out


def read_cancellations(combined_xlsx_path, excluded_plans=None):
    """
    Read every Cancelled row from the Name Tracking tab of the Combined File.
    Returns a list of dicts: {display_name, plan, start_date, end_date}.
    Rows whose plan is in `excluded_plans` are skipped.
    """
    excluded_plans = excluded_plans or set()
    wb = load_workbook(combined_xlsx_path, data_only=True)
    if "Name Tracking" not in wb.sheetnames:
        return []
    ws = wb["Name Tracking"]
    _, col_index, nt_rows = read_name_tracking(ws)
    out = []
    for r in nt_rows:
        if r["Status"] != "Cancelled":
            continue
        if r["Plan"] in excluded_plans:
            continue
        if not r["End Date"]:
            continue
        out.append({
            "display_name": r["Display Name"],
            "plan": r["Plan"],
            "start_date": r["Start Date"],
            "end_date": r["End Date"],
        })
    return out


RUN_HISTORY_HEADER = ["Pull Date", "New Count", "Cancelled Count",
                      "New Members", "Cancelled Members"]


def append_run_history(wb, pull_date, cancelled_this_run, new_this_run):
    """Append a summary row to the Run History tab (creates it on first run)."""
    if "Run History" not in wb.sheetnames:
        ws = wb.create_sheet(title="Run History", index=0)
        ws.append(RUN_HISTORY_HEADER)
    else:
        ws = wb["Run History"]
    new_names = "; ".join("%s (%s)" % (r["display_name"], r["plan"]) for r in new_this_run)
    cx_names = "; ".join("%s (%s)" % (r["display_name"], r["plan"]) for r in cancelled_this_run)
    ws.append([pull_date, len(new_this_run), len(cancelled_this_run), new_names, cx_names])


def append_snapshot_tab(wb, pull_date, csv_rows):
    """Create a new tab named pull_date (with " (N)" suffix on collision) and fill it."""
    base = pull_date
    name = base
    n = 2
    while name in wb.sheetnames:
        name = "%s (%d)" % (base, n)
        n += 1
    ws = wb.create_sheet(title=name)
    ws.append(SNAPSHOT_HEADER)
    for r in csv_rows:
        ws.append([_norm(r.get(col)) for col in SNAPSHOT_HEADER])
    return name


def prune_old_snapshot_tabs(wb, keep_last=7):
    """
    Delete dated snapshot tabs older than the most recent `keep_last`.
    Preserves Run History, Name Tracking, Sheet1, and any non-date-named tab.
    Snapshot tabs match YYYY-MM-DD (with optional " (N)" suffix).
    Keeps memory down and the file small — Run History captures the audit trail
    for older days, so old snapshots are safe to drop.
    """
    import re as _re
    date_pat = _re.compile(r"^\d{4}-\d{2}-\d{2}( \(\d+\))?$")
    snapshot_tabs = [name for name in wb.sheetnames if date_pat.match(name)]
    # Sort by the date prefix ascending; keep the newest `keep_last`
    snapshot_tabs.sort(key=lambda s: s[:10])
    to_delete = snapshot_tabs[:-keep_last] if len(snapshot_tabs) > keep_last else []
    for name in to_delete:
        del wb[name]
    return to_delete


def update_combined_file(combined_xlsx_path, csv_rows, pull_date, output_xlsx_path, excluded_plans=None):
    """
    Main entry. Loads the .xlsx, computes the diff, updates Name Tracking in place,
    appends a new snapshot tab (with the FULL CSV — excluded plans included so the
    snapshot stays raw), saves to output_xlsx_path, and returns a diff report.

    `excluded_plans` is a set of plan names to ignore for the Name Tracking diff
    only. Cancelled/added flags will not fire for these plans, and existing rows
    in Name Tracking with these plans are left untouched.
    """
    excluded_plans = excluded_plans or set()
    wb = load_workbook(combined_xlsx_path)
    if "Name Tracking" not in wb.sheetnames:
        raise RuntimeError("'Name Tracking' tab not found in Combined File.")

    nt = wb["Name Tracking"]
    _, col_index, nt_rows = read_name_tracking(nt)

    current_active = build_current_active(csv_rows, excluded_plans=excluded_plans)
    # "Present in CSV in any non-terminal status" — Active / Draft / Sent / etc.
    # A member showing up here still exists in HCP; we should NOT mark them
    # Cancelled just because their status flipped from Active to Draft. The
    # active-count metric still only counts Active, but Name Tracking treats
    # them as still-alive.
    terminal_statuses = {"Cancelled", "Expired"}
    csv_present = set()
    for r in csv_rows:
        plan = _norm(r.get("Plan"))
        if plan in excluded_plans:
            continue
        status = _norm(r.get("Status"))
        if status in terminal_statuses:
            continue  # counts as a cancellation, not "still present"
        name = _norm(r.get("Display Name"))
        if name and plan:
            csv_present.add((name, plan))

    nt_all_keys = {(r["Display Name"], r["Plan"]) for r in nt_rows}
    # Skip Name Tracking rows whose plan is excluded — don't cancel them.
    nt_active_rows = [r for r in nt_rows
                      if r["Status"] == "Active" and r["Plan"] not in excluded_plans]

    # --- Cancellations: prior-Active NT keys that are truly gone from CSV
    #     (absent entirely, or present with terminal status). Drafts/Sents
    #     are protective — they leave NT Active. ---
    cancelled_this_run = []
    for r in nt_active_rows:
        key = (r["Display Name"], r["Plan"])
        if key not in csv_present:
            # Update in place
            nt.cell(row=r["_row"], column=col_index["Status"]).value = "Cancelled"
            nt.cell(row=r["_row"], column=col_index["End Date"]).value = pull_date
            cancelled_this_run.append({
                "display_name": r["Display Name"],
                "plan": r["Plan"],
                "start_date": r["Start Date"],
                "end_date": pull_date,
                "first_seen": r["First Seen (Pull)"],
            })

    # --- New: current-active keys not previously tracked at all ---
    new_this_run = []
    # append at bottom
    next_row = nt.max_row + 1
    for key, info in current_active.items():
        if key in nt_all_keys:
            continue
        name, plan = key
        nt.cell(row=next_row, column=col_index["Display Name"]).value = name
        nt.cell(row=next_row, column=col_index["Plan"]).value = plan
        nt.cell(row=next_row, column=col_index["Start Date"]).value = info["start_date"]
        nt.cell(row=next_row, column=col_index["Status"]).value = "Active"
        nt.cell(row=next_row, column=col_index["End Date"]).value = None
        nt.cell(row=next_row, column=col_index["First Seen (Pull)"]).value = pull_date
        new_this_run.append({
            "display_name": name,
            "plan": plan,
            "start_date": info["start_date"],
            "first_seen": pull_date,
        })
        next_row += 1

    # --- Append the new snapshot tab ---
    tab_name = append_snapshot_tab(wb, pull_date, csv_rows)

    # --- Log this run to the Run History tab ---
    append_run_history(wb, pull_date, cancelled_this_run, new_this_run)

    # --- Prune old snapshot tabs so the file stays small (daily cadence) ---
    pruned = prune_old_snapshot_tabs(wb, keep_last=7)
    if pruned:
        print("Pruned %d old snapshot tabs: %s" % (len(pruned), ", ".join(pruned)))

    wb.save(output_xlsx_path)

    return {
        "pull_date": pull_date,
        "new_tab_name": tab_name,
        "prior_active_count": len(nt_active_rows),
        "current_active_count": len(current_active),
        "cancelled_this_run": cancelled_this_run,
        "new_this_run": new_this_run,
        "output_path": output_xlsx_path,
    }
